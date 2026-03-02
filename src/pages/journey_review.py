"""
Worker's Journey 보정 리뷰 페이지.

작업자를 선택하면:
  - 원본(Raw) Journey 타임라인
  - 보정(Corrected) Journey 타임라인
  - 보정된 행 상세 비교 테이블
  - 토글 스위치로 보정 로직 설명 확인
  - CSV / Excel 다운로드 (원본 + 원본_장소 + 보정_장소 열 포함)
"""
from __future__ import annotations

import io

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from src.data.schema import RawColumns, ProcessedColumns
from src.utils.theme import Color, apply_theme
from src.utils.place_utils import sort_places_by_similarity, sort_places_smart
from src.utils.constants import (
    ACTIVE_RATIO_ZERO_THRESHOLD,
    LOCATION_SMOOTHING_WINDOW,
    HELMET_RACK_MIN_DURATION_MIN,
    COORD_OUTLIER_THRESHOLD,
    NIGHT_HOURS_START,
    DAWN_HOURS_END,
    LUNCH_START,
    LUNCH_END,
    ACTIVITY_COLORS,
    ACTIVITY_LABELS,
)


# ─── 보정 유형 정의 ────────────────────────────────────────────────────
CORRECTION_TYPES = {
    "helmet_rack":   {"label": "헬멧 거치 통일",   "color": "#E74C3C", "icon": "🪖"},
    "noise_smooth":  {"label": "이동 노이즈 제거",  "color": "#F39C12", "icon": "🌊"},
    "coord_fix":     {"label": "좌표 이상치 보정",  "color": "#9B59B6", "icon": "📍"},
}


def render(df: pd.DataFrame) -> None:
    if df is None or df.empty:
        st.warning("데이터 없음")
        return

    # ── 헤더 ───────────────────────────────────────────────────────
    st.markdown("""
    <div class="kpi-banner">
        <h1>🔧 Worker's Journey 보정 리뷰</h1>
        <p>원본 데이터와 보정 데이터를 비교하여 Journey 보정 결과를 검증합니다.</p>
    </div>""", unsafe_allow_html=True)

    # ── 작업자 선택 ────────────────────────────────────────────────
    worker_keys = sorted(df[ProcessedColumns.WORKER_KEY].unique())
    worker_label = {}
    for wk in worker_keys:
        sub = df[df[ProcessedColumns.WORKER_KEY] == wk]
        name    = sub[RawColumns.WORKER].iloc[0]
        company = sub[RawColumns.COMPANY].iloc[0]
        n_corr  = sub[ProcessedColumns.IS_CORRECTED].sum()
        badge   = f"  ✦ {n_corr}건 보정" if n_corr > 0 else ""
        worker_label[wk] = f"{name}  ({company}){badge}"

    col_sel, col_stat = st.columns([1, 2])
    with col_sel:
        selected_key = st.selectbox(
            "👷 작업자 선택",
            options=worker_keys,
            format_func=lambda k: worker_label.get(k, k),
        )

    wdf = df[df[ProcessedColumns.WORKER_KEY] == selected_key].copy()
    wdf = wdf.sort_values(RawColumns.TIME).reset_index(drop=True)

    name    = wdf[RawColumns.WORKER].iloc[0]
    company = wdf[RawColumns.COMPANY].iloc[0]
    tag_id  = wdf[RawColumns.TAG].iloc[0]

    stats = _compute_correction_stats(wdf)

    with col_stat:
        _render_worker_stat_bar(name, company, tag_id, stats)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 다운로드 섹션 ───────────────────────────────────────────────
    _render_download_section(wdf, name)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 보정 로직 토글 ─────────────────────────────────────────────
    show_logic = st.toggle(
        "🔍 보정 로직 보기",
        value=False,
        help="적용된 보정 알고리즘과 이 작업자 데이터에서의 실제 예시를 표시합니다.",
    )
    if show_logic:
        _render_correction_logic(wdf)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── 메인 탭 ────────────────────────────────────────────────────
    tab0, tab1, tab2, tab3 = st.tabs([
        "🗺️ 전체 Journey",
        "📊 보정 비교 (원본 vs 보정)",
        "📋 보정 상세 테이블",
        "⚡ 활성비율 비교",
    ])

    with tab0:
        _render_full_journey_overview(wdf, name, stats, df)

    with tab1:
        _render_journey_comparison(wdf, name, df)

    with tab2:
        _render_correction_table(wdf, selected_key)

    with tab3:
        _render_active_ratio_comparison(wdf, name)


# ── 다운로드: 내보내기 DataFrame 생성 ────────────────────────────────

# raw 출력 시 포함할 원본 컬럼 순서 (CSV 원본과 동일)
_RAW_EXPORT_COLS = [
    RawColumns.TIME,
    RawColumns.WORKER,
    RawColumns.ZONE,
    RawColumns.BUILDING,
    RawColumns.FLOOR,
    RawColumns.PLACE,
    RawColumns.X,
    RawColumns.Y,
    RawColumns.TAG,
    RawColumns.TAG_TYPE,
    RawColumns.COMPANY,
    RawColumns.EQUIPMENT,
    RawColumns.SIGNAL_COUNT,
    RawColumns.ACTIVE_SIGNAL_COUNT,
]

# 다운로드 파일에 추가할 비교 컬럼 이름
_COL_RAW_PLACE  = "원본_장소"
_COL_CORR_PLACE = "보정_장소"
_COL_CHANGED    = "장소변경여부"


def _build_export_df(wdf: pd.DataFrame) -> pd.DataFrame:
    """
    원본 raw 컬럼 + [원본_장소 / 보정_장소 / 장소변경여부] 3열을 오른쪽에 붙인
    내보내기용 DataFrame을 반환한다.
    """
    available = [c for c in _RAW_EXPORT_COLS if c in wdf.columns]
    export = wdf[available].copy()

    # 우측에 비교 열 추가
    export[_COL_RAW_PLACE]  = wdf[RawColumns.PLACE].values
    export[_COL_CORR_PLACE] = wdf[ProcessedColumns.CORRECTED_PLACE].values
    export[_COL_CHANGED]    = (
        wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]
    ).map({True: "Y", False: ""}).values

    # 날짜/시간 포맷을 원본 CSV와 동일하게 (YYYY.MM.DD HH:MM:SS)
    export[RawColumns.TIME] = pd.to_datetime(
        export[RawColumns.TIME]
    ).dt.strftime("%Y.%m.%d %H:%M:%S")

    return export


def _to_csv_bytes(export_df: pd.DataFrame) -> bytes:
    """UTF-8 BOM CSV bytes 반환 (Excel에서 한글 깨짐 방지)."""
    return export_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def _to_excel_bytes(export_df: pd.DataFrame) -> bytes:
    """
    Excel(.xlsx) bytes 반환.
    - 헤더 행: 진한 네이비 배경 + 흰 글자
    - 장소변경여부 = "Y"인 행: 연한 주황 배경 (보정된 행 강조)
    - 원본_장소 / 보정_장소 / 장소변경여부 3열: 연한 파랑 배경으로 구분
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # openpyxl 없으면 단순 bytes로 반환
        buf = io.BytesIO()
        export_df.to_excel(buf, index=False, engine=None)
        return buf.getvalue()

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Journey 비교"

    # ── 스타일 정의 ──────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1B3A6B")   # 네이비
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    corr_fill    = PatternFill("solid", fgColor="FEF5E7")   # 연한 주황 (보정행)
    extra_fill   = PatternFill("solid", fgColor="EBF5FB")   # 연한 파랑 (비교열)
    changed_fill = PatternFill("solid", fgColor="FDEBD0")   # 진한 주황 (변경행)
    thin_border  = Border(
        left=Side(style="thin", color="D0D7E8"),
        right=Side(style="thin", color="D0D7E8"),
        top=Side(style="thin", color="D0D7E8"),
        bottom=Side(style="thin", color="D0D7E8"),
    )

    cols        = list(export_df.columns)
    extra_start = cols.index(_COL_RAW_PLACE) + 1   # 1-based
    changed_col = cols.index(_COL_CHANGED) + 1

    # ── 헤더 행 ──────────────────────────────────────────────────
    for ci, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    ws.row_dimensions[1].height = 28

    # ── 데이터 행 ─────────────────────────────────────────────────
    for ri, (_, row) in enumerate(export_df.iterrows(), start=2):
        is_changed = str(row.get(_COL_CHANGED, "")) == "Y"
        for ci, col_name in enumerate(cols, start=1):
            val  = row[col_name]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
            # 우선순위: 변경행 > 비교열 > 일반
            if ci >= extra_start:
                cell.fill = changed_fill if is_changed else extra_fill
            elif is_changed:
                cell.fill = corr_fill

    # ── 열 너비 자동 조정 ────────────────────────────────────────
    for ci, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(ci)
        # 헤더 길이 기준으로 최소 너비 산정
        header_len = len(str(col_name))
        max_data   = export_df[col_name].astype(str).str.len().max() if len(export_df) else 0
        width      = max(header_len, min(max_data, 30)) + 2
        ws.column_dimensions[col_letter].width = width

    # ── 첫 행 고정 (스크롤 시 헤더 유지) ────────────────────────
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _render_download_section(wdf: pd.DataFrame, name: str) -> None:
    """CSV / Excel 다운로드 버튼 영역을 렌더링한다."""
    export_df = _build_export_df(wdf)

    # 파일명 기준: 날짜 추출
    try:
        date_tag = pd.to_datetime(wdf[RawColumns.TIME].iloc[0]).strftime("%Y%m%d")
    except Exception:
        date_tag = "unknown"

    # 작업자 이름의 특수문자 제거 (파일명 안전 처리)
    safe_name = name.replace("*", "X").replace(" ", "_").replace("/", "_")
    base_name = f"journey_{safe_name}_{date_tag}"

    diff_n = (export_df[_COL_CHANGED] == "Y").sum()

    st.markdown(f"""
    <div style="background:#F8F9FB;border:1px solid #E0E7F0;border-radius:12px;
                padding:0.9rem 1.3rem;display:flex;align-items:center;gap:1rem;
                flex-wrap:wrap;">
        <div style="font-size:0.88rem;font-weight:600;color:{Color.TEXT_DARK};flex:1 1 200px;">
            📥 &nbsp;Journey 데이터 내보내기
            <span style="font-size:0.8rem;font-weight:400;color:{Color.TEXT_MUTED};margin-left:0.5rem;">
                {len(export_df):,}행 · 장소변경 <b style="color:{Color.ACCENT};">{diff_n}건</b>
                &nbsp;· 원본_장소 / 보정_장소 열 포함
            </span>
        </div>
    </div>""", unsafe_allow_html=True)

    dl_col1, dl_col2, _ = st.columns([1, 1, 3])

    with dl_col1:
        csv_bytes = _to_csv_bytes(export_df)
        st.download_button(
            label="⬇ CSV 다운로드",
            data=csv_bytes,
            file_name=f"{base_name}.csv",
            mime="text/csv",
            use_container_width=True,
            help="UTF-8 BOM 인코딩 · Excel에서 한글 정상 표시",
        )

    with dl_col2:
        excel_bytes = _to_excel_bytes(export_df)
        st.download_button(
            label="⬇ Excel 다운로드",
            data=excel_bytes,
            file_name=f"{base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="보정된 행은 주황색으로 강조 표시됩니다",
        )


# ── 컴포넌트: 작업자 상태바 ──────────────────────────────────────────

def _compute_correction_stats(wdf: pd.DataFrame) -> dict:
    """보정 통계를 3가지 카테고리로 계산.

    Returns:
        total_n         : 전체 행 수 (= 1분 단위 총 기록 시간)
        n_place_changed : 장소명 실제 변경 (원본 장소 ≠ 보정 장소)
        n_coord_only    : 좌표만 보정 (IS_CORRECTED=True 이지만 장소명 동일)
        n_unchanged     : 원본 유지 (IS_CORRECTED=False)
        place_change_rate : n_place_changed / total_n
    """
    total_n     = len(wdf)
    is_corr     = wdf[ProcessedColumns.IS_CORRECTED] == True
    place_diff  = wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]
    n_place_changed = int(place_diff.sum())
    n_coord_only    = int((is_corr & ~place_diff).sum())
    n_unchanged     = int((~is_corr).sum())
    place_change_rate = n_place_changed / total_n if total_n > 0 else 0.0
    return {
        "total_n": total_n,
        "n_place_changed": n_place_changed,
        "n_coord_only": n_coord_only,
        "n_unchanged": n_unchanged,
        "place_change_rate": place_change_rate,
    }


def _render_worker_stat_bar(
    name: str, company: str, tag_id: str,
    stats: dict,
) -> None:
    """3-카테고리 보정 통계 바를 렌더링."""
    total_n          = stats["total_n"]
    n_place_changed  = stats["n_place_changed"]
    n_coord_only     = stats["n_coord_only"]
    n_unchanged      = stats["n_unchanged"]
    place_change_rate = stats["place_change_rate"]

    has_corr = n_place_changed > 0 or n_coord_only > 0
    border_color = Color.ACCENT if has_corr else Color.SAFE

    st.markdown(f"""
    <div style="background:#FFFFFF;border:1px solid #E8ECF4;border-radius:12px;
                padding:1rem 1.4rem;display:flex;gap:2rem;align-items:center;
                border-left:5px solid {border_color};">
        <div>
            <div style="font-size:1.1rem;font-weight:700;color:{Color.PRIMARY};">{name}</div>
            <div style="font-size:0.8rem;color:{Color.TEXT_MUTED};">{company} · {tag_id}</div>
        </div>
        <div style="display:flex;gap:1.6rem;margin-left:auto;flex-wrap:wrap;">
            <div style="text-align:center;">
                <div style="font-size:1.35rem;font-weight:700;color:{Color.PRIMARY};">{total_n:,}</div>
                <div style="font-size:0.7rem;color:{Color.TEXT_MUTED};">전체 기록(분)</div>
            </div>
            <div style="text-align:center;">
                <div style="font-size:1.35rem;font-weight:700;color:{Color.DANGER if n_place_changed>0 else Color.SAFE};">{n_place_changed:,}</div>
                <div style="font-size:0.7rem;color:{Color.TEXT_MUTED};">장소명 변경</div>
            </div>
            <div style="text-align:center;">
                <div style="font-size:1.35rem;font-weight:700;color:#9B59B6;">{n_coord_only:,}</div>
                <div style="font-size:0.7rem;color:{Color.TEXT_MUTED};">좌표만 보정</div>
            </div>
            <div style="text-align:center;">
                <div style="font-size:1.35rem;font-weight:700;color:{Color.SAFE};">{n_unchanged:,}</div>
                <div style="font-size:0.7rem;color:{Color.TEXT_MUTED};">원본 유지</div>
            </div>
            <div style="text-align:center;">
                <div style="font-size:1.35rem;font-weight:700;color:{Color.ACCENT};">{place_change_rate:.1%}</div>
                <div style="font-size:0.7rem;color:{Color.TEXT_MUTED};">장소 변경률</div>
            </div>
        </div>
    </div>""", unsafe_allow_html=True)

    # 파라미터 정의 토글
    with st.expander("📖 파라미터 정의 보기", expanded=False):
        st.markdown(f"""
        | 파라미터 | 정의 | 값 |
        |---|---|---|
        | **전체 기록(분)** | 작업자의 분 단위 위치 기록 총 수. 하루 최대 1,440분 | `{total_n:,}분` |
        | **장소명 변경** | 보정 전후 장소명이 실제로 달라진 행 수.<br>헬멧 거치 통일 또는 이동 노이즈 제거로 발생 | `{n_place_changed:,}건` |
        | **좌표만 보정** | 장소명은 그대로이나 X·Y 좌표가 보정된 행 수.<br>보정 블록 내 대표 좌표로 통일된 경우 | `{n_coord_only:,}건` |
        | **원본 유지** | 보정 알고리즘이 적용되지 않은 행 수 (`보정여부=False`) | `{n_unchanged:,}건` |
        | **장소 변경률** | 장소명 변경 건수 ÷ 전체 기록 수.<br>실제 의미 있는 보정이 얼마나 이루어졌는지를 나타냄 | `{place_change_rate:.1%}` |

        > ⚠️ **주의**: `보정여부=True`인 행(보정 블록에 포함된 행) = 장소명 변경({n_place_changed:,}) + 좌표만 보정({n_coord_only:,}) = `{n_place_changed+n_coord_only:,}건`.
        > 이 숫자를 "보정률"로 사용하면 실제보다 과장될 수 있습니다.
        """, unsafe_allow_html=True)


# ── 컴포넌트: 전체 Journey 통합 뷰 ──────────────────────────────────

def _get_global_axes_jr(df: pd.DataFrame, full_df: pd.DataFrame = None, use_original: bool = False) -> tuple:
    """
    당일 전체 데이터 기준 공통 축 (x=0~24h, y=전체 장소).
    
    ★ 핵심: Y축에 전체 데이터(full_df)의 모든 장소를 포함하여 작업자 간 일관된 비교 가능.
    
    Args:
        df: 현재 작업자 데이터
        full_df: 전체 데이터 (모든 작업자 포함 — Y축 장소 목록 + 이동 기반 정렬에 사용)
        use_original: 원본 장소 사용 여부
    """
    if df is None or df.empty:
        return pd.Timestamp("2000-01-01 00:00:00"), pd.Timestamp("2000-01-02 00:00:00"), []

    date_val = df[ProcessedColumns.DATE].iloc[0] if ProcessedColumns.DATE in df.columns else df[RawColumns.TIME].iloc[0]
    if hasattr(date_val, "strftime"):
        date_str = date_val.strftime("%Y-%m-%d")
    else:
        s = str(date_val).replace("-", "")[:8]
        date_str = f"{s[:4]}-{s[4:6]}-{s[6:8]}" if len(s) >= 8 else "2000-01-01"

    t_min = pd.Timestamp(f"{date_str} 00:00:00")
    t_max = t_min + pd.Timedelta(days=1)

    place_col = RawColumns.PLACE if use_original else ProcessedColumns.CORRECTED_PLACE
    
    # ★ 전체 데이터(full_df)의 모든 장소 수집 — 작업자 간 일관된 Y축
    place_set: set = set()
    data_for_places = full_df if full_df is not None and not full_df.empty else df
    
    for col in [RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE]:
        if col in data_for_places.columns:
            for p in data_for_places[col].astype(str):
                if p and p != "nan":
                    place_set.add(p)
    
    # 이동 기반 스마트 정렬 (전체 데이터 활용)
    transition_df = full_df if full_df is not None else df
    place_order = sort_places_smart(list(place_set), transition_df, place_col)
    return t_min, t_max, place_order


def _render_full_journey_overview(wdf: pd.DataFrame, name: str, stats: dict, df: pd.DataFrame) -> None:
    """전체 Journey 통합 뷰. 메인: 원본 장소 + 보정 핀 오버레이."""
    sorted_df = wdf.sort_values(RawColumns.TIME).reset_index(drop=True)
    t_min, t_max, all_places = _get_global_axes_jr(wdf, full_df=df, use_original=True)

    # Gantt 블록 생성 (원본 장소 기준 — 전체 Journey)
    gantt_rows: list = []
    current_place: str | None = None
    start_ts = None
    block_rows: list = []

    for _, row in sorted_df.iterrows():
        place = str(row.get(RawColumns.PLACE, "Unknown"))
        ts    = row[RawColumns.TIME]
        if place != current_place:
            if current_place and block_rows:
                _flush_gantt_block(gantt_rows, current_place, start_ts, block_rows)
            current_place, start_ts, block_rows = place, ts, [row]
        else:
            block_rows.append(row)
    if current_place and block_rows:
        _flush_gantt_block(gantt_rows, current_place, start_ts, block_rows)

    if not gantt_rows:
        st.info("Journey 데이터 없음")
        return

    gantt_df = pd.DataFrame(gantt_rows)
    t_data_min = sorted_df[RawColumns.TIME].min()
    t_data_max = sorted_df[RawColumns.TIME].max()
    t_min_str = t_data_min.strftime("%H:%M")
    t_max_str = t_data_max.strftime("%H:%M")

    n_places    = len(gantt_df["장소"].unique())
    n_blocks    = len(gantt_df)
    n_corrected = stats["n_place_changed"]

    # ── 요약 KPI ────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("기록 시간", f"{t_min_str} ~ {t_max_str}",
              help=f"실제 기록 범위. 총 {len(sorted_df)}분")
    c2.metric("방문 장소 수", f"{n_places}개",
              help="보정 후 기준으로 이동한 서로 다른 장소의 수")
    c3.metric("장소 블록 수", f"{n_blocks}건",
              help="같은 장소에 연속으로 머문 구간 수")
    c4.metric(
        "장소명 보정", f"{n_corrected}건" if n_corrected > 0 else "없음",
        delta="장소명 변경됨" if n_corrected > 0 else None,
        delta_color="inverse" if n_corrected > 0 else "off",
        help="보정 알고리즘으로 장소명이 실제 변경된 행 수 (좌표 보정 제외)",
    )

    st.markdown("<hr style='margin:0.6rem 0;border-color:#E0E7F0'>", unsafe_allow_html=True)

    # ── 전체 Journey Gantt ───────────────────────────────────────
    st.markdown(
        "##### 📍 전체 Journey &nbsp;"
        "<small style='color:#888;font-weight:400'>x축: 00:00~24:00 · y축: 당일 전체 장소 (비교 기준 통일)</small>",
        unsafe_allow_html=True,
    )
    st.caption(f"📅 이 작업자 실제 기록: {t_min_str} ~ {t_max_str} ({len(sorted_df)}분)")
    st.caption("📌 위 차트: 원본 Journey. ✓ 마커 = 보정된 구간 (마우스 오버 시 원본→보정 확인)")

    fig = go.Figure()
    shown_cats: set = set()

    for _, row in gantt_df.iterrows():
        act       = row.get("활동상태", "off_duty")
        bar_color = ACTIVITY_COLORS.get(act, "#B0B8C8")
        act_label = ACTIVITY_LABELS.get(act, act)
        dur_ms    = (row["종료"] - row["시작"]).total_seconds() * 1000
        show_leg  = act not in shown_cats
        shown_cats.add(act)

        fig.add_trace(go.Bar(
            base=row["시작"].strftime("%Y-%m-%d %H:%M:%S"),
            x=[dur_ms],
            y=[row["장소"]],
            orientation="h",
            marker_color=bar_color,
            marker_line=dict(color="white", width=0.5),
            opacity=0.92,
            name=act_label,
            legendgroup=act,
            showlegend=show_leg,
            customdata=[[
                row["평균활성비율"], row["체류(분)"], act_label,
                row["고활성(분)"], row["저활성(분)"], row["장소유형"],
            ]],
            hovertemplate=(
                "<b>%{y}</b><br>"
                "체류: %{customdata[1]}분<br>"
                "활성비율: %{customdata[0]:.1%}<br>"
                "상태: %{customdata[2]}<br>"
                "고활성: %{customdata[3]}분 · 저활성: %{customdata[4]}분<br>"
                "장소유형: %{customdata[5]}"
                "<extra></extra>"
            ),
        ))

    # 보정 핀 (원본 Journey 위에 오버레이 — 원본 장소 행에 표시)
    if n_corrected > 0:
        place_diff_mask = wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]
        corr_rows  = wdf[place_diff_mask].sort_values(RawColumns.TIME)
        leg_added  = False
        for _, row in corr_rows.iterrows():
            ts_ms      = row[RawColumns.TIME].strftime("%Y-%m-%d %H:%M:%S")
            orig_place = str(row[RawColumns.PLACE])
            corr_place = str(row[ProcessedColumns.CORRECTED_PLACE])
            time_str   = row[RawColumns.TIME].strftime("%H:%M")
            ratio      = float(row.get(ProcessedColumns.ACTIVE_RATIO, 0.0))
            fig.add_trace(go.Bar(
                base=ts_ms,
                x=[60_000],
                y=[orig_place],
                orientation="h",
                marker_color=Color.ACCENT,
                marker_line=dict(color="#C0550A", width=2),
                opacity=1.0,
                name="🔧 보정 포인트",
                legendgroup="correction_pin",
                showlegend=not leg_added,
                customdata=[[orig_place, corr_place, ratio, time_str]],
                hovertemplate=(
                    "<b>%{customdata[3]} · 🔧 보정됨</b><br>"
                    "원본: <b>%{customdata[0]}</b><br>"
                    "→ 보정: <b>%{customdata[1]}</b><br>"
                    "활성비율: %{customdata[2]:.1%}"
                    "<extra>✦ 보정</extra>"
                ),
            ))
            leg_added = True

    chart_h = max(380, len(all_places) * 30 + 120) if all_places else max(380, n_places * 30 + 120)
    fig = apply_theme(fig, "", height=chart_h)
    fig.update_layout(
        barmode="overlay",
        bargap=0.28,
        legend=dict(
            orientation="h", yanchor="bottom", y=1.01,
            xanchor="left", x=0, font=dict(size=10),
        ),
        xaxis=dict(
            type="date", tickformat="%H:%M", showgrid=True,
            gridcolor="#E4EAF4", tickfont=dict(size=10, color="#3A4A6A"),
            range=[t_min.isoformat(), t_max.isoformat()],
            dtick=2 * 60 * 60 * 1000,  # 2시간 간격
            tickmode="linear",
            tick0=t_min.replace(hour=(t_min.hour // 2) * 2, minute=0, second=0).isoformat(),
        ),
        yaxis=dict(
            title_text="장소",
            **(dict(categoryorder="array", categoryarray=all_places) if all_places else dict(categoryorder="category descending")),
            tickfont=dict(size=9, color="#1B2A4A"),
        ),
        margin=dict(l=0, r=10, t=40, b=0),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── 보정된 사항 (전체 Journey 아래) ─────────────────────────────
    if n_corrected > 0:
        st.markdown("##### 🔧 보정된 사항")
        st.caption("위 원본 Journey 상에서 보정된 구간입니다. 보정이 합리적인지 확인하세요.")
        corr_detail = wdf[place_diff_mask][[RawColumns.TIME, RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE, ProcessedColumns.ACTIVE_RATIO]].copy()
        corr_detail = corr_detail.sort_values(RawColumns.TIME)
        corr_detail["시간"] = corr_detail[RawColumns.TIME].dt.strftime("%H:%M")
        corr_detail["원본 장소"] = corr_detail[RawColumns.PLACE]
        corr_detail["보정 장소"] = corr_detail[ProcessedColumns.CORRECTED_PLACE]
        corr_detail["활성비율"] = corr_detail[ProcessedColumns.ACTIVE_RATIO].apply(lambda x: f"{x:.1%}")
        disp = corr_detail[["시간", "원본 장소", "보정 장소", "활성비율"]]
        st.dataframe(disp, use_container_width=True, hide_index=True)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── 활성비율 시계열 ──────────────────────────────────────────
    st.markdown("##### 📈 활성비율 시계열")
    act_fig = go.Figure()
    act_fig.add_trace(go.Scatter(
        x=sorted_df[RawColumns.TIME],
        y=sorted_df[ProcessedColumns.ACTIVE_RATIO],
        mode="lines",
        name="활성비율",
        line=dict(color="#2E6FD9", width=1.5),
        fill="tozeroy",
        fillcolor="rgba(46,111,217,0.10)",
        hovertemplate="%{x|%H:%M}<br>활성비율: %{y:.1%}<extra></extra>",
    ))

    if n_corrected > 0:
        place_diff_mask = wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]
        corr_df = sorted_df[sorted_df[RawColumns.PLACE] != sorted_df[ProcessedColumns.CORRECTED_PLACE]]
        act_fig.add_trace(go.Scatter(
            x=corr_df[RawColumns.TIME],
            y=corr_df[ProcessedColumns.ACTIVE_RATIO],
            mode="markers",
            name="🔧 보정 포인트",
            marker=dict(color=Color.ACCENT, size=6, symbol="diamond",
                        line=dict(color="#C0550A", width=1)),
            hovertemplate="%{x|%H:%M}<br>보정 포인트<br>활성비율: %{y:.1%}<extra></extra>",
        ))

    act_fig = apply_theme(act_fig, "", height=180)
    act_fig.update_layout(
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0, font=dict(size=10)),
        margin=dict(l=0, r=10, t=30, b=0),
    )
    act_fig.update_xaxes(
        type="date", tickformat="%H:%M", showgrid=True,
        range=[t_min.isoformat(), t_max.isoformat()],
        dtick=2 * 60 * 60 * 1000,  # 2시간 간격
        tickmode="linear",
        tick0=t_min.replace(hour=(t_min.hour // 2) * 2, minute=0, second=0).isoformat(),
    )
    act_fig.update_yaxes(range=[-0.05, 1.05], tickformat=".0%", title_text="활성비율")
    st.plotly_chart(act_fig, use_container_width=True)

    # ── Journey 상세 기록 (x축=시간, y축=장소 Gantt) ─────────────────
    st.markdown("##### 📋 Journey 상세 기록 <small style='color:#888;font-weight:400'>(x축: 시간 · y축: 전체 장소)</small>",
                unsafe_allow_html=True)
    if n_corrected > 0:
        st.caption("🔶 주황 핀 = 장소명 보정된 구간")

    detail_fig = go.Figure()
    shown_cats_detail: set = set()

    for _, row in gantt_df.iterrows():
        act       = row.get("활동상태", "off_duty")
        bar_color = ACTIVITY_COLORS.get(act, "#B0B8C8")
        act_label = ACTIVITY_LABELS.get(act, act)
        dur_ms    = (row["종료"] - row["시작"]).total_seconds() * 1000
        show_leg  = act not in shown_cats_detail
        shown_cats_detail.add(act)

        detail_fig.add_trace(go.Bar(
            base=row["시작"].strftime("%Y-%m-%d %H:%M:%S"),
            x=[dur_ms],
            y=[row["장소"]],
            orientation="h",
            marker_color=bar_color,
            marker_line=dict(color="white", width=0.5),
            opacity=0.92,
            name=act_label,
            legendgroup=act,
            showlegend=show_leg,
            customdata=[[
                row["평균활성비율"], row["체류(분)"],
                row["시작"].strftime("%H:%M"),
                (row["종료"] - pd.Timedelta(minutes=1)).strftime("%H:%M"),
                row["고활성(분)"], row["저활성(분)"], act_label,
            ]],
            hovertemplate=(
                "<b>%{y}</b><br>%{customdata[2]} ~ %{customdata[3]}<br>"
                "체류: %{customdata[1]}분 · 활성비율: %{customdata[0]:.1%}<br>"
                "상태: %{customdata[6]}<br>"
                "고활성: %{customdata[4]}분 · 저활성: %{customdata[5]}분<extra></extra>"
            ),
        ))

    if n_corrected > 0:
        corr_rows = wdf[wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]].sort_values(RawColumns.TIME)
        leg_added = False
        for _, row in corr_rows.iterrows():
            ts_ms      = row[RawColumns.TIME].strftime("%Y-%m-%d %H:%M:%S")
            orig_place = str(row[RawColumns.PLACE])
            corr_place = str(row[ProcessedColumns.CORRECTED_PLACE])
            time_str   = row[RawColumns.TIME].strftime("%H:%M")
            ratio      = float(row.get(ProcessedColumns.ACTIVE_RATIO, 0.0))
            detail_fig.add_trace(go.Bar(
                base=ts_ms,
                x=[60_000],
                y=[orig_place],
                orientation="h",
                marker_color=Color.ACCENT,
                marker_line=dict(color="#C0550A", width=2),
                opacity=1.0,
                name="🔧 보정",
                legendgroup="corr_detail",
                showlegend=not leg_added,
                customdata=[[orig_place, corr_place, ratio, time_str]],
                hovertemplate=(
                    "<b>%{customdata[3]} · 보정</b><br>"
                    "원본: %{customdata[0]}<br>→ 보정: %{customdata[1]}<br>"
                    "활성비율: %{customdata[2]:.1%}<extra></extra>"
                ),
            ))
            leg_added = True

    detail_h = max(320, len(all_places) * 36 + 100) if all_places else max(320, n_places * 36 + 100)
    detail_fig = apply_theme(detail_fig, "", height=detail_h)
    detail_fig.update_layout(
        barmode="overlay",
        bargap=0.28,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0, font=dict(size=10)),
        xaxis=dict(
            type="date", tickformat="%H:%M", title_text="시간",
            showgrid=True, gridcolor="#E4EAF4",
            range=[t_min.isoformat(), t_max.isoformat()],
            dtick=2 * 60 * 60 * 1000,  # 2시간 간격
            tickmode="linear",
            tick0=t_min.replace(hour=(t_min.hour // 2) * 2, minute=0, second=0).isoformat(),
        ),
        yaxis=dict(
            title_text="장소",
            **(dict(categoryorder="array", categoryarray=all_places) if all_places else dict(categoryorder="category descending")),
            tickfont=dict(size=9, color="#1B2A4A"),
        ),
        margin=dict(l=0, r=10, t=40, b=40),
    )
    st.plotly_chart(detail_fig, use_container_width=True)

    with st.expander("📄 블록별 상세 테이블 보기", expanded=False):
        narrative_rows: list = []
        for blk in gantt_df.to_dict("records"):
            blk_mask = (
                (sorted_df[RawColumns.TIME] >= blk["시작"])
                & (sorted_df[RawColumns.TIME] <  blk["종료"])
            )
            blk_df     = sorted_df[blk_mask]
            n_corr_blk = int((blk_df[RawColumns.PLACE] != blk_df[ProcessedColumns.CORRECTED_PLACE]).sum())
            orig_set   = set(blk_df.loc[blk_df[RawColumns.PLACE] != blk_df[ProcessedColumns.CORRECTED_PLACE], RawColumns.PLACE].tolist())
            narrative_rows.append({
                "시작": blk["시작"].strftime("%H:%M"),
                "종료": (blk["종료"] - pd.Timedelta(minutes=1)).strftime("%H:%M"),
                "장소": blk["장소"],
                "장소유형": blk["장소유형"],
                "체류(분)": blk["체류(분)"],
                "활성비율": f"{blk['평균활성비율']:.1%}",
                "보정(건)": n_corr_blk,
                "원본": ", ".join(sorted(orig_set)) if orig_set else "-",
            })
        st.dataframe(pd.DataFrame(narrative_rows), use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 장소별 누적 체류 시간 ─────────────────────────────────────
    st.markdown("##### 🏢 장소별 누적 체류 시간 (상위 15개)")
    top_dwell = gantt_df.groupby("장소")["체류(분)"].sum().nlargest(15).reset_index()
    dwell_fig = go.Figure(go.Bar(
        x=top_dwell["체류(분)"],
        y=top_dwell["장소"],
        orientation="h",
        marker_color="#2E6FD9",
        marker_line_color="white",
        text=top_dwell["체류(분)"].apply(lambda v: f"{v}분"),
        textposition="outside",
        hovertemplate="<b>%{y}</b><br>%{x}분<extra></extra>",
    ))
    dwell_fig = apply_theme(dwell_fig, "", height=max(260, len(top_dwell) * 24 + 80))
    dwell_fig.update_xaxes(title_text="체류 시간 (분)")
    dwell_fig.update_yaxes(autorange="reversed")
    dwell_fig.update_layout(margin=dict(l=0, r=60, t=20, b=0))
    st.plotly_chart(dwell_fig, use_container_width=True)


def _flush_gantt_block(rows: list, place: str, start_ts, buf: list) -> None:
    """Journey Review용 Gantt 블록 flush 헬퍼."""
    from src.utils.constants import WORK_INTENSITY_HIGH_THRESHOLD, WORK_INTENSITY_LOW_THRESHOLD
    from src.utils.place_classifier import classify_block_activity
    end_ts     = buf[-1][RawColumns.TIME] + pd.Timedelta(minutes=1)
    ratios     = [r.get(ProcessedColumns.ACTIVE_RATIO, 0) for r in buf]
    avg_ratio  = sum(ratios) / len(ratios)
    high_min   = sum(1 for r in ratios if r >= WORK_INTENSITY_HIGH_THRESHOLD)
    low_min    = sum(1 for r in ratios if WORK_INTENSITY_LOW_THRESHOLD <= r < WORK_INTENSITY_HIGH_THRESHOLD)
    place_type = buf[0].get(ProcessedColumns.PLACE_TYPE, "UNKNOWN")
    
    # state_detail에서 transit_arrival 확인 (전환 이동 태깅 반영)
    state_details = [r.get(ProcessedColumns.STATE_DETAIL, "") for r in buf]
    transit_arrival_count = sum(1 for s in state_details if s == "transit_arrival")
    
    # 블록의 과반이 transit_arrival이면 transit으로 표시
    if transit_arrival_count > len(buf) / 2:
        activity = "transit"
    else:
        activity = classify_block_activity(place_type, avg_ratio, start_ts.hour)
    
    rows.append({
        "장소":        place,
        "시작":        start_ts,
        "종료":        end_ts,
        "장소유형":    place_type,
        "활동상태":    activity,
        "평균활성비율": round(avg_ratio, 3),
        "체류(분)":    len(buf),
        "고활성(분)":  high_min,
        "저활성(분)":  low_min,
    })


# ── 컴포넌트: 보정 로직 설명 토글 ────────────────────────────────────

def _render_correction_logic(wdf: pd.DataFrame) -> None:
    """보정에 사용된 알고리즘과 실제 적용 예시를 표시한다."""

    # 실제 적용 여부 판단
    diff_mask = wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]
    has_diff  = diff_mask.any()

    # 야간 비활성 패턴 존재 여부
    night_inactive = (
        (wdf[ProcessedColumns.HOUR].apply(lambda h: h >= NIGHT_HOURS_START or h < DAWN_HOURS_END))
        & (wdf[ProcessedColumns.ACTIVE_RATIO] <= ACTIVE_RATIO_ZERO_THRESHOLD)
    )
    lunch_inactive = (
        (wdf[ProcessedColumns.HOUR].apply(lambda h: LUNCH_START <= h < LUNCH_END))
        & (wdf[ProcessedColumns.ACTIVE_RATIO] <= ACTIVE_RATIO_ZERO_THRESHOLD)
    )

    st.markdown("""
    <div style="background:#F8F9FB;border:1px solid #E0E7F0;border-radius:14px;padding:1.4rem 1.8rem;">
        <div style="font-size:1rem;font-weight:700;color:{primary};margin-bottom:1rem;">
            🔬 적용된 보정 로직
        </div>
    """.replace("{primary}", Color.PRIMARY), unsafe_allow_html=True)

    # ── 로직 1: 헬멧 거치 장소 통일 ─────────────────────────────
    with st.expander("🪖 Rule 1 — 헬멧 거치 장소 통일 (야간 / 점심 비활성 패턴)", expanded=True):
        st.markdown(f"""
        **적용 조건**
        - 야간/새벽 (`{NIGHT_HOURS_START}시 이후` 또는 `{DAWN_HOURS_END}시 이전`) + 활성비율 ≤ {ACTIVE_RATIO_ZERO_THRESHOLD}
        - 또는 점심시간 (`{LUNCH_START}시 ~ {LUNCH_END}시`) + 활성비율 ≤ {ACTIVE_RATIO_ZERO_THRESHOLD}
        - 연속 구간 ≥ **{HELMET_RACK_MIN_DURATION_MIN}분** 이상 지속

        **보정 내용**
        - 해당 구간에서 가장 많이 등장한 헬멧 걸이대 장소로 **모든 행 통일**
        - 근처 ±10행 내 헬멧 걸이대가 없으면 보정 미적용
        """)

        night_n  = night_inactive.sum()
        lunch_n  = lunch_inactive.sum()

        c1, c2 = st.columns(2)
        c1.metric("야간 비활성 구간", f"{night_n}분", help="야간/새벽 + 활성비율=0인 행 수")
        c2.metric("점심 비활성 구간", f"{lunch_n}분", help="점심시간 + 활성비율=0인 행 수")

        # 실제 보정 예시
        rule1_diff = wdf[diff_mask & night_inactive | diff_mask & lunch_inactive]
        if not rule1_diff.empty:
            st.markdown("**📌 실제 보정 예시 (상위 10건)**")
            ex = rule1_diff[
                [RawColumns.TIME, RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE,
                 ProcessedColumns.ACTIVE_RATIO, ProcessedColumns.HOUR]
            ].head(10).copy()
            ex[RawColumns.TIME] = ex[RawColumns.TIME].dt.strftime("%H:%M")
            ex[ProcessedColumns.ACTIVE_RATIO] = ex[ProcessedColumns.ACTIVE_RATIO].apply(lambda v: f"{v:.2f}")
            ex.columns = ["시간", "원본 장소", "보정 장소", "활성비율", "시"]
            st.dataframe(ex, use_container_width=True, hide_index=True)
        else:
            st.info("이 작업자에게는 Rule 1이 적용된 행이 없습니다.")

    # ── 로직 2: 이동 노이즈 제거 ─────────────────────────────────
    with st.expander(f"🌊 Rule 2 — 이동 노이즈 제거 (슬라이딩 윈도우 최빈값, window={LOCATION_SMOOTHING_WINDOW})"):
        st.markdown(f"""
        **적용 조건**
        - 슬라이딩 윈도우 크기: **{LOCATION_SMOOTHING_WINDOW}분** (앞뒤 ±{LOCATION_SMOOTHING_WINDOW//2}행)
        - 윈도우 내 최빈값(Mode)이 현재 장소와 다를 경우 보정 적용
        - 이미 Rule 1로 보정된 행은 제외

        **목적**
        - `A, A, A, B, A, B, B, B` → `A, A, A, A, B, B, B, B` (실제 이동으로 스무딩)
        - 위치 측정 오차로 인한 단발성 장소 튀김 제거
        """)

        # 노이즈 보정 예시: 원본과 보정이 다르지만 야간/점심이 아닌 경우
        noise_diff = wdf[diff_mask & ~night_inactive & ~lunch_inactive]
        if not noise_diff.empty:
            st.markdown(f"**📌 노이즈 보정 예시 ({len(noise_diff)}건)**")
            ex2 = noise_diff[
                [RawColumns.TIME, RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE,
                 ProcessedColumns.ACTIVE_RATIO]
            ].head(10).copy()
            ex2[RawColumns.TIME] = ex2[RawColumns.TIME].dt.strftime("%H:%M")
            ex2[ProcessedColumns.ACTIVE_RATIO] = ex2[ProcessedColumns.ACTIVE_RATIO].apply(lambda v: f"{v:.2f}")
            ex2.columns = ["시간", "원본 장소", "보정 장소", "활성비율"]
            st.dataframe(ex2, use_container_width=True, hide_index=True)
        else:
            st.info("이 작업자에게는 Rule 2가 적용된 행이 없습니다.")

    # ── 로직 3: 좌표 이상치 보정 ──────────────────────────────────
    with st.expander(f"📍 Rule 3 — 좌표 이상치 보정 (임계값: {COORD_OUTLIER_THRESHOLD}px)"):
        st.markdown(f"""
        **적용 조건**
        - 같은 건물+층(위치키) 내에서 연속된 좌표 간 변화량 > **{COORD_OUTLIER_THRESHOLD}픽셀**
        - 실외(OUTDOOR)와 실내(INDOOR) 좌표계는 **독립적으로** 처리

        **보정 내용**
        - 이상치로 판단된 좌표를 `NaN`으로 마스킹
        - 선형 보간(Linear Interpolation)으로 대체
        """)

        # 좌표 보정 확인
        x_diff = (wdf[RawColumns.X] - wdf[ProcessedColumns.CORRECTED_X]).abs()
        y_diff = (wdf[RawColumns.Y] - wdf[ProcessedColumns.CORRECTED_Y]).abs()
        coord_fixed = ((x_diff > 1) | (y_diff > 1)).sum()

        c1, c2 = st.columns(2)
        c1.metric("X 좌표 보정", f"{(x_diff > 1).sum()}건")
        c2.metric("Y 좌표 보정", f"{(y_diff > 1).sum()}건")

        if coord_fixed > 0:
            st.markdown(f"**📌 좌표 보정 샘플 ({coord_fixed}건)**")
            coord_ex = wdf[(x_diff > 1) | (y_diff > 1)][
                [RawColumns.TIME, ProcessedColumns.LOCATION_KEY,
                 RawColumns.X, ProcessedColumns.CORRECTED_X,
                 RawColumns.Y, ProcessedColumns.CORRECTED_Y]
            ].head(10).copy()
            coord_ex[RawColumns.TIME] = coord_ex[RawColumns.TIME].dt.strftime("%H:%M")
            coord_ex.columns = ["시간", "위치키", "원본X", "보정X", "원본Y", "보정Y"]
            st.dataframe(coord_ex, use_container_width=True, hide_index=True)
        else:
            st.info("이 작업자에게는 좌표 보정이 적용되지 않았습니다.")

    st.markdown("</div>", unsafe_allow_html=True)


# ── 컴포넌트: Journey 비교 차트 ───────────────────────────────────────

def _render_journey_comparison(wdf: pd.DataFrame, name: str, df: pd.DataFrame) -> None:
    """원본 Journey와 보정 Journey 타임라인을 위아래로 나란히 표시.
    축: x=0~24시, y=당일 전체 장소 (비교 기준 통일)
    """

    orig_gantt = _build_gantt(wdf, RawColumns.PLACE)
    corr_gantt = _build_gantt(wdf, ProcessedColumns.CORRECTED_PLACE)

    if orig_gantt.empty and corr_gantt.empty:
        st.info("Journey 데이터 없음")
        return

    diff_n = (wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]).sum()

    t_min, t_max, all_places = _get_global_axes_jr(wdf, full_df=df)
    x_range = [t_min, t_max]

    # 보정 없는 경우 안내
    if diff_n == 0:
        st.success("✅ 원본과 보정 Journey가 동일합니다. 보정이 적용되지 않았습니다.")
        st.markdown("**📍 Journey Timeline**")
        n_y = len(all_places) if all_places else len(orig_gantt["장소"].unique())
        fig = _make_gantt_figure(orig_gantt, f"{name} · Journey", height=max(280, n_y * 34 + 80), x_range=x_range, y_places=all_places)
        st.plotly_chart(fig, use_container_width=True)
        return

    # ── 원본 / 보정 타임라인 ────────────────────────────────────
    st.markdown(f"""
    <div style="background:#FEF5E7;border:1px solid {Color.ACCENT};border-radius:10px;
                padding:0.7rem 1rem;margin-bottom:1rem;font-size:0.88rem;color:{Color.TEXT_DARK};">
        ⚠️ &nbsp;<b>{diff_n}개 행</b>의 장소명이 보정되었습니다.
        &nbsp;<b style="color:{Color.ACCENT};">◆ 주황 핀(1분)</b>이 정확한 보정 포인트입니다.
        원본 차트: 원래 위치 표시 / 보정 차트: 이동된 위치 표시
        &nbsp;|&nbsp; x축: 00:00~24:00 · y축: 당일 전체 장소 (비교 기준 통일)
    </div>""", unsafe_allow_html=True)

    n_y = len(all_places) if all_places else max(len(orig_gantt["장소"].unique()), len(corr_gantt["장소"].unique()))
    chart_h = max(280, n_y * 34 + 80)

    col_l, col_r = st.columns(2, gap="medium")

    with col_l:
        st.markdown(f"""
        <div style="text-align:center;background:{Color.BG_MUTED};border-radius:8px;
                    padding:0.5rem;margin-bottom:0.5rem;font-weight:600;
                    color:{Color.TEXT_DARK};font-size:0.9rem;">
            📂 원본 Journey (Raw)
        </div>""", unsafe_allow_html=True)
        fig_orig = _make_gantt_figure(orig_gantt, "", height=chart_h, x_range=x_range, y_places=all_places)
        fig_orig = _add_correction_pins(fig_orig, wdf, RawColumns.PLACE)
        st.plotly_chart(fig_orig, use_container_width=True)

    with col_r:
        st.markdown(f"""
        <div style="text-align:center;background:#EAF4EA;border-radius:8px;
                    padding:0.5rem;margin-bottom:0.5rem;font-weight:600;
                    color:{Color.SAFE};font-size:0.9rem;">
            ✅ 보정 Journey (Corrected)
        </div>""", unsafe_allow_html=True)

        fig_corr = _make_gantt_figure(corr_gantt, "", height=chart_h, x_range=x_range, y_places=all_places)
        fig_corr = _add_correction_pins(fig_corr, wdf, ProcessedColumns.CORRECTED_PLACE)
        st.plotly_chart(fig_corr, use_container_width=True)

    # ── 보정 구간 변화 상세 ────────────────────────────────────────
    st.markdown("**🔄 장소별 보정 전→후 변화 요약**")
    _render_correction_change_summary(wdf)


def _render_correction_change_summary(wdf: pd.DataFrame) -> None:
    """원본 장소 → 보정 장소 변화 패턴 집계 및 시각화."""
    diff_df = wdf[wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]].copy()
    if diff_df.empty:
        return

    change_count = (
        diff_df.groupby([RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE])
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
    )

    fig = go.Figure()
    for i, row in change_count.iterrows():
        fig.add_trace(go.Bar(
            x=[row["count"]],
            y=[f"{row[RawColumns.PLACE][:20]}  →  {row[ProcessedColumns.CORRECTED_PLACE][:20]}"],
            orientation="h",
            marker_color=Color.ACCENT,
            text=f"{row['count']}건",
            textposition="outside",
            showlegend=False,
            hovertemplate=f"원본: <b>{row[RawColumns.PLACE]}</b><br>"
                          f"보정: <b>{row[ProcessedColumns.CORRECTED_PLACE]}</b><br>"
                          f"건수: {row['count']}건<extra></extra>",
        ))
    fig = apply_theme(fig, "보정 패턴 (원본 → 보정)", height=max(200, len(change_count) * 40 + 80))
    fig.update_yaxes(autorange="reversed")
    st.plotly_chart(fig, use_container_width=True)


# ── 컴포넌트: 보정 상세 테이블 ───────────────────────────────────────

def _render_correction_table(wdf: pd.DataFrame, worker_key: str = "") -> None:
    """원본 vs 보정 차이가 있는 행의 상세 테이블."""
    diff_df = wdf[wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]].copy()

    if diff_df.empty:
        st.success("✅ 이 작업자는 장소 변경이 없습니다.")
        st.info(f"보정여부=True 행 수: {wdf[ProcessedColumns.IS_CORRECTED].sum()}건 "
                f"(좌표 보정만 적용, 장소명 변경 없음)")
        return

    # ── 요약 배너 ──────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:#FEF9EF;border:1px solid {Color.ACCENT};border-radius:10px;
                padding:0.6rem 1rem;margin-bottom:0.8rem;font-size:0.88rem;">
        📝 &nbsp;총 <b style="color:{Color.ACCENT};">{len(diff_df):,}건</b>의 장소 변경이 있습니다.
        아래 표에서 원본 장소와 보정 장소를 행별로 확인하세요.
    </div>""", unsafe_allow_html=True)

    # ── 필터 (worker_key를 key에 포함해 작업자 변경 시 선택 초기화) ──
    wk_hash = abs(hash(worker_key)) % 100000
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        hour_range = st.slider(
            "시간대 필터",
            0, 23, (0, 23),
            key=f"corr_table_hour_{wk_hash}",
        )
    with col_f2:
        orig_places = ["전체"] + sorted(diff_df[RawColumns.PLACE].unique().tolist())
        sel_place   = st.selectbox(
            "원본 장소 필터", orig_places,
            index=0,
            key=f"corr_table_place_{wk_hash}",
        )

    # ── 필터 적용 ──────────────────────────────────────────────
    mask = (
        (diff_df[ProcessedColumns.HOUR] >= hour_range[0])
        & (diff_df[ProcessedColumns.HOUR] <= hour_range[1])
    )
    if sel_place != "전체":
        mask &= diff_df[RawColumns.PLACE] == sel_place

    filtered = diff_df[mask].copy()

    st.caption(f"필터 결과: {len(filtered):,}건 / 전체 {len(diff_df):,}건")

    if filtered.empty:
        st.warning("필터 조건에 맞는 데이터가 없습니다. 필터를 조정해 주세요.")
        return

    # ── 테이블 컬럼 구성 ────────────────────────────────────────
    # 존재하는 컬럼만 선택
    desired = [
        RawColumns.TIME,
        ProcessedColumns.HOUR,
        RawColumns.PLACE,
        ProcessedColumns.CORRECTED_PLACE,
        ProcessedColumns.ACTIVE_RATIO,
        RawColumns.SIGNAL_COUNT,
        RawColumns.ACTIVE_SIGNAL_COUNT,
        ProcessedColumns.PERIOD_TYPE,
    ]
    avail = [c for c in desired if c in filtered.columns]
    disp = filtered[avail].copy()

    if RawColumns.TIME in disp.columns:
        disp[RawColumns.TIME] = pd.to_datetime(disp[RawColumns.TIME]).dt.strftime("%H:%M")
    if ProcessedColumns.ACTIVE_RATIO in disp.columns:
        disp[ProcessedColumns.ACTIVE_RATIO] = disp[ProcessedColumns.ACTIVE_RATIO].apply(
            lambda v: f"{v:.2f}"
        )

    col_labels = {
        RawColumns.TIME:                   "시간",
        ProcessedColumns.HOUR:             "시",
        RawColumns.PLACE:                  "원본 장소",
        ProcessedColumns.CORRECTED_PLACE:  "보정 장소",
        ProcessedColumns.ACTIVE_RATIO:     "활성비율",
        RawColumns.SIGNAL_COUNT:           "신호수",
        RawColumns.ACTIVE_SIGNAL_COUNT:    "활성신호",
        ProcessedColumns.PERIOD_TYPE:      "활동유형",
    }
    disp = disp.rename(columns={k: v for k, v in col_labels.items() if k in disp.columns})

    table_h = min(520, max(120, len(filtered) * 35 + 42))
    st.dataframe(disp, use_container_width=True, hide_index=True, height=table_h)


# ── 컴포넌트: 활성비율 비교 ─────────────────────────────────────────

def _render_active_ratio_comparison(wdf: pd.DataFrame, name: str) -> None:
    """활성비율 시계열 + 보정 구간 오버레이."""
    sorted_df = wdf.sort_values(RawColumns.TIME)
    diff_mask = sorted_df[RawColumns.PLACE] != sorted_df[ProcessedColumns.CORRECTED_PLACE]

    fig = go.Figure()

    # 활성비율 면적
    fig.add_trace(go.Scatter(
        x=sorted_df[RawColumns.TIME],
        y=sorted_df[ProcessedColumns.ACTIVE_RATIO],
        mode="lines",
        name="활성비율",
        line=dict(color=Color.SECONDARY, width=1.5),
        fill="tozeroy",
        fillcolor="rgba(46,111,217,0.10)",
        hovertemplate="%{x|%H:%M} · 활성비율: %{y:.1%}<extra></extra>",
    ))

    # 보정된 시점 마커 (주황 다이아몬드)
    corr_df = sorted_df[diff_mask]
    if not corr_df.empty:
        fig.add_trace(go.Scatter(
            x=corr_df[RawColumns.TIME],
            y=corr_df[ProcessedColumns.ACTIVE_RATIO],
            mode="markers",
            name="보정 적용 시점",
            marker=dict(
                color=Color.ACCENT,
                size=8,
                symbol="diamond",
                line=dict(color=Color.PRIMARY, width=1),
            ),
            customdata=corr_df[[RawColumns.PLACE, ProcessedColumns.CORRECTED_PLACE]].values,
            hovertemplate=(
                "%{x|%H:%M}<br>"
                "원본: <b>%{customdata[0]}</b><br>"
                "보정: <b>%{customdata[1]}</b><extra></extra>"
            ),
        ))

    # 기준선
    fig.add_hline(y=0.3, line_dash="dash", line_color=Color.SAFE, line_width=1.5,
                  annotation_text="작업 기준(0.3)", annotation_font_size=10)
    fig.add_hline(y=0.05, line_dash="dot", line_color=Color.WARNING, line_width=1,
                  annotation_text="비활동(0.05)", annotation_font_size=10)

    fig = apply_theme(fig, f"{name} · 활성비율 시계열 (◆ 보정 시점)", height=350)
    fig.update_xaxes(tickformat="%H:%M", title_text="시간")
    fig.update_yaxes(range=[-0.05, 1.05], title_text="활성비율")
    st.plotly_chart(fig, use_container_width=True)

    # 원본/보정 장소 텍스트 비교 (시간축 산점도)
    st.markdown("**📍 원본 vs 보정 장소 산점도**")
    fig2 = go.Figure()

    # 원본 (회색)
    fig2.add_trace(go.Scatter(
        x=sorted_df[RawColumns.TIME],
        y=sorted_df[RawColumns.PLACE],
        mode="markers",
        name="원본",
        marker=dict(color="#C8D6E8", size=5, symbol="circle"),
        hovertemplate="%{x|%H:%M} · %{y}<extra>원본</extra>",
    ))

    # 보정된 행만 오버레이 (주황)
    if not corr_df.empty:
        fig2.add_trace(go.Scatter(
            x=corr_df[RawColumns.TIME],
            y=corr_df[ProcessedColumns.CORRECTED_PLACE],
            mode="markers",
            name="보정 후",
            marker=dict(
                color=Color.ACCENT,
                size=10,
                symbol="diamond",
                line=dict(color=Color.PRIMARY, width=1.2),
            ),
            hovertemplate="%{x|%H:%M} · 보정: %{y}<extra>보정</extra>",
        ))

    fig2 = apply_theme(fig2, "원본(회색 ●) vs 보정(주황 ◆) 장소 비교", height=max(300, len(sorted_df[RawColumns.PLACE].unique()) * 22 + 100))
    fig2.update_xaxes(tickformat="%H:%M", title_text="시간")
    fig2.update_yaxes(title_text="장소")
    st.plotly_chart(fig2, use_container_width=True)


# ── 내부 헬퍼 ─────────────────────────────────────────────────────────

def _build_gantt(df: pd.DataFrame, place_col: str, max_gap_min: int = 5) -> pd.DataFrame:
    """
    장소 컬럼 기준으로 연속 구간 Gantt 데이터 생성.
    
    Args:
        df: 작업자 Journey DataFrame
        place_col: 장소 컬럼명
        max_gap_min: 최대 허용 공백 (분). 이 이상 공백이면 별도 블록으로 분리.
    """
    # 시간순 정렬 필수
    sorted_df = df.sort_values(RawColumns.TIME).reset_index(drop=True)
    
    rows, cur_place, start_ts, buf, prev_ts = [], None, None, [], None

    for _, row in sorted_df.iterrows():
        place = str(row.get(place_col, "Unknown"))
        ts    = row[RawColumns.TIME]
        
        # 시간 공백 체크: 이전 데이터와 max_gap_min 이상 차이나면 블록 종료
        if prev_ts is not None and cur_place and buf:
            gap_min = (ts - prev_ts).total_seconds() / 60
            if gap_min > max_gap_min:
                _flush(rows, cur_place, start_ts, buf)
                cur_place = None
                start_ts = None
                buf = []
        
        if place != cur_place:
            if cur_place and buf:
                _flush(rows, cur_place, start_ts, buf)
            cur_place = place
            start_ts  = ts
            buf       = [row]
        else:
            buf.append(row)
        
        prev_ts = ts
        
    if cur_place and buf:
        _flush(rows, cur_place, start_ts, buf)

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def _flush(rows: list, place: str, start_ts, buf: list) -> None:
    from src.utils.constants import WORK_INTENSITY_HIGH_THRESHOLD, WORK_INTENSITY_LOW_THRESHOLD
    from src.utils.place_classifier import classify_block_activity
    end_ts     = buf[-1][RawColumns.TIME] + pd.Timedelta(minutes=1)
    ratios     = [r.get(ProcessedColumns.ACTIVE_RATIO, 0) for r in buf]
    avg_ratio  = sum(ratios) / len(ratios)
    high_min   = sum(1 for r in ratios if r >= WORK_INTENSITY_HIGH_THRESHOLD)
    low_min    = sum(1 for r in ratios if WORK_INTENSITY_LOW_THRESHOLD <= r < WORK_INTENSITY_HIGH_THRESHOLD)
    place_type = buf[0].get(ProcessedColumns.PLACE_TYPE, "UNKNOWN")
    
    # state_detail에서 transit_arrival 확인 (전환 이동 태깅 반영)
    state_details = [r.get(ProcessedColumns.STATE_DETAIL, "") for r in buf]
    transit_arrival_count = sum(1 for s in state_details if s == "transit_arrival")
    
    # 블록의 과반이 transit_arrival이면 transit으로 표시
    if transit_arrival_count > len(buf) / 2:
        activity = "transit"
    else:
        activity = classify_block_activity(place_type, avg_ratio, start_ts.hour)
    
    rows.append({
        "장소":        place,
        "시작":        start_ts,
        "종료":        end_ts,
        "장소유형":    place_type,
        "활동상태":    activity,
        "평균활성비율": round(avg_ratio, 3),
        "체류(분)":    len(buf),
        "고활성(분)":  high_min,
        "저활성(분)":  low_min,
    })


_ONE_MIN_MS = 60_000  # 1분 = 60,000 ms


def _add_correction_pins(
    fig: go.Figure,
    wdf: pd.DataFrame,
    place_col: str,
) -> go.Figure:
    """
    보정된 행 각각에 정확한 1분 너비의 주황 마커 바를 오버레이한다.

    place_col: 핀을 꽂을 Y축 장소 컬럼
      - RawColumns.PLACE               → 원본 장소 위치에 표시
      - ProcessedColumns.CORRECTED_PLACE → 보정 장소 위치에 표시
    """
    diff_rows = wdf[wdf[RawColumns.PLACE] != wdf[ProcessedColumns.CORRECTED_PLACE]]
    if diff_rows.empty:
        return fig

    legend_added = False
    for _, row in diff_rows.iterrows():
        ts_ms      = row[RawColumns.TIME].strftime("%Y-%m-%d %H:%M:%S")
        y_place    = str(row[place_col])
        orig_place = str(row[RawColumns.PLACE])
        corr_place = str(row[ProcessedColumns.CORRECTED_PLACE])
        time_str   = row[RawColumns.TIME].strftime("%H:%M")
        ratio      = row.get(ProcessedColumns.ACTIVE_RATIO, 0.0)

        fig.add_trace(go.Bar(
            base=ts_ms,
            x=[_ONE_MIN_MS],
            y=[y_place],
            orientation="h",
            marker_color=Color.ACCENT,          # 주황
            marker_line=dict(color="#C0550A", width=2),
            opacity=1.0,
            name="보정 포인트",
            legendgroup="correction_pin",
            showlegend=not legend_added,
            customdata=[[orig_place, corr_place, ratio, time_str]],
            hovertemplate=(
                "<b>%{customdata[3]} · 보정 포인트</b><br>"
                "원본 장소: <b>%{customdata[0]}</b><br>"
                "보정 장소: <b>%{customdata[1]}</b><br>"
                "활성비율: %{customdata[2]:.1%}"
                "<extra>✦ 보정</extra>"
            ),
        ))
        legend_added = True

    return fig


def _make_gantt_figure(
    gantt_df: pd.DataFrame,
    title: str,
    height: int = 320,
    x_range: list | None = None,
    y_places: list | None = None,
) -> go.Figure:
    """Gantt DataFrame으로 Plotly 수평 막대 차트 생성.

    x_range: [start_ts, end_ts] — 0~24시 등 공통 x축.
    y_places: 당일 전체 장소 목록 — 비교용 공통 y축.
    """
    if gantt_df.empty:
        return go.Figure()

    fig = go.Figure()

    shown_cats: set = set()

    for _, row in gantt_df.iterrows():
        act       = row.get("활동상태", "off_duty")
        bar_color = ACTIVITY_COLORS.get(act, "#B0B8C8")
        act_label = ACTIVITY_LABELS.get(act, act)
        dur_ms    = (row["종료"] - row["시작"]).total_seconds() * 1000
        show_leg  = act not in shown_cats
        shown_cats.add(act)

        fig.add_trace(go.Bar(
            base=row["시작"].strftime("%Y-%m-%d %H:%M:%S"),
            x=[dur_ms],
            y=[row["장소"]],
            orientation="h",
            marker_color=bar_color,
            marker_line=dict(color="white", width=0.5),
            opacity=0.92,
            name=act_label,
            legendgroup=act,
            showlegend=show_leg,
            customdata=[[row["평균활성비율"], row["체류(분)"], row.get("고활성(분)", 0), row.get("저활성(분)", 0), act_label]],
            hovertemplate=(
                "<b>%{y}</b><br>"
                "체류: %{customdata[1]}분<br>"
                "활성비율: %{customdata[0]:.1%}<br>"
                "상태: %{customdata[4]}<br>"
                "고활성: %{customdata[2]}분 · 저활성: %{customdata[3]}분"
                "<extra></extra>"
            ),
        ))

    xaxis_cfg: dict = dict(
        type="date",
        tickformat="%H:%M",
        showgrid=True,
        gridcolor="#E4EAF4",
        tickfont=dict(size=10, color="#3A4A6A"),
        dtick=2 * 60 * 60 * 1000,  # 2시간 간격
        tickmode="linear",
    )
    if x_range:
        xaxis_cfg["range"] = [
            x_range[0].isoformat(),
            x_range[1].isoformat(),
        ]
        # tick0 설정
        tick0_hour = (x_range[0].hour // 2) * 2
        tick0_time = x_range[0].replace(hour=tick0_hour, minute=0, second=0)
        xaxis_cfg["tick0"] = tick0_time.isoformat()

    fig = apply_theme(fig, title, height=height)
    fig.update_layout(
        barmode="overlay",
        bargap=0.28,
        legend=dict(
            orientation="h",
            yanchor="bottom", y=1.01,
            xanchor="left", x=0,
            font=dict(size=10),
        ),
        xaxis=xaxis_cfg,
        yaxis=dict(
            title_text="장소",
            **(dict(categoryorder="array", categoryarray=y_places) if y_places else dict(categoryorder="category descending")),
            tickfont=dict(size=10, color="#1B2A4A"),
        ),
    )
    return fig
